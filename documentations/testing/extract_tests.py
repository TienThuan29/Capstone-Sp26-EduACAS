#!/usr/bin/env python3
"""Extract UnitTest data from Excel to Markdown."""

import openpyxl

wb = openpyxl.load_workbook(
    '/home/tienthuan29/workspaces/projects/Capstone-Sp26-EduACAS/documentations/testing/UnitTest_Function1_Function26.xlsx',
    data_only=True
)

out_lines = []

def emit(line):
    out_lines.append(line)


def parse_sheet(fn):
    sheet_name = f'Function {fn}'
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    # ── Header info ──────────────────────────────────────────────
    code = f"F{fn:03d}"
    func_name = ""
    created_by = ""
    executed_by = ""
    loc = ""
    test_req = ""

    for row in all_rows[:8]:
        for i, cell in enumerate(row):
            if cell == 'Function Code' and i + 2 < len(row):
                code = str(row[i + 2])
            if cell == 'Function Name' and i + 6 < len(row) and row[i + 6]:
                func_name = str(row[i + 6]).strip()
            if cell == 'Created By' and i + 2 < len(row) and row[i + 2]:
                created_by = str(row[i + 2]).strip()
            if cell == 'Executed By' and i + 6 < len(row) and row[i + 6]:
                executed_by = str(row[i + 6]).strip()
            if cell == 'Lines  of code' and i + 2 < len(row) and row[i + 2]:
                loc = str(row[i + 2]).strip()
            if cell == 'Test requirement' and i + 2 < len(row) and row[i + 2]:
                test_req = str(row[i + 2]).strip()

    # ── Find UTCD column indices ──────────────────────────────────
    tc_cols = []
    utcd_ids = []
    for row in all_rows:
        for i, cell in enumerate(row):
            if cell and isinstance(cell, str) and cell.startswith('UTCD-'):
                if i not in tc_cols:
                    tc_cols.append(i)
                    utcd_ids.append(cell)
    n_tc = len(utcd_ids)

    def o_tcs(row):
        return [tc_cols.index(i) for i, c in enumerate(row) if c == 'O' and i in tc_cols]

    # ── Data buckets ─────────────────────────────────────────────
    precond_items  = []     # [(desc, set_of_tcs)]
    param_data     = {}      # { param_name: { tc_idx: desc } }
    confirm_items  = []
    return_items   = []
    except_items   = []
    log_items     = []
    tc_types      = {i: "" for i in range(n_tc)}
    tc_pf         = {i: "" for i in range(n_tc)}

    # Section states
    S_PRECOND  = "PRECOND"
    S_PARAM    = "PARAM"
    S_CONFIRM  = "CONFIRM"
    S_RETURN   = "RETURN"
    S_EXCEPT   = "EXCEPTION"
    S_LOG      = "LOG"

    section = S_PRECOND
    current_param = None   # current parameter name from col_B

    for row_idx, row in enumerate(all_rows):
        if row_idx < 9:
            continue
        if len(row) < 4:
            continue

        col_a = str(row[0]).strip() if row[0] else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        col_c = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        col_d = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        def o_tcs_local():
            return [tc_cols.index(i) for i, c in enumerate(row) if c == 'O' and i in tc_cols]

        # ── Skip rows with no descriptive data ───────────────────
        skip_vals = ('Passed/Failed', 'Executed Date', 'Defect ID',
                     'Type(N : Normal, A : Abnormal, B : Boundary)',
                     'Confirm', 'Return', 'Exception', 'Log message', 'Condition')
        if col_d in skip_vals:
            continue

        # ── Section transitions ────────────────────────────────────
        if col_a == 'Condition' and col_b == 'Precondition ':
            section = S_PRECOND
            current_param = None
            continue

        if col_a == 'Confirm':
            section = S_CONFIRM
            current_param = None
            continue

        if col_b == 'Return' or (col_a == 'Confirm' and col_b == ''):
            section = S_RETURN
            current_param = None
            continue

        if col_b == 'Exception':
            section = S_EXCEPT
            current_param = None
            continue

        if col_b == 'Log message':
            section = S_LOG
            current_param = None
            continue

        # ── Type row ─────────────────────────────────────────────
        if 'Type(N : Normal' in col_b:
            for tc_idx, col_num in enumerate(tc_cols):
                if col_num < len(row) and row[col_num]:
                    tc_types[tc_idx] = str(row[col_num]).strip()
            continue

        if col_a == 'Result' and col_b == 'Passed/Failed':
            for tc_idx, col_num in enumerate(tc_cols):
                if col_num < len(row) and row[col_num]:
                    tc_pf[tc_idx] = str(row[col_num]).strip()
            continue

        # ── col_B param name row (param section marker, no col_D) ──
        # Row has col_B filled (param name) but col_D empty.
        # This sets current_param but has no O markers.
        if col_b and not col_d:
            section = S_PARAM
            current_param = col_b
            continue

        # Skip rows with no data
        if not col_d:
            continue

        tcs = o_tcs_local()

        # ── col_C param name row (has both col_C and col_D) ──────
        # col_C has the param name, col_D has the first value
        if col_c and col_c != "None":
            section = S_PARAM
            current_param = col_c
            if tcs:
                param_data.setdefault(col_c, {})[tcs[0]] = col_d
                for tc in tcs[1:]:
                    param_data.setdefault(col_c, {})[tc] = col_d

        # ── Continuation/value row (col_C empty, has col_D + O markers) ─
        elif not col_c and tcs:
            if section == S_PRECOND:
                precond_items.append((col_d, set(tcs)))
            elif section == S_PARAM and current_param:
                for tc in tcs:
                    param_data.setdefault(current_param, {})[tc] = col_d
            elif section == S_CONFIRM:
                confirm_items.append((col_d, set(tcs)))
            elif section == S_RETURN:
                return_items.append((col_d, set(tcs)))
            elif section == S_EXCEPT:
                except_items.append((col_d, set(tcs)))
            elif section == S_LOG:
                log_items.append((col_d, set(tcs)))

        # ── Continuation/value row without O markers (e.g. col_C="None") ──
        elif not col_c and not tcs:
            # These are the col_C="None" rows - they have a description but no O markers
            # They describe preconditions (e.g. "input is a string")
            if section == S_PRECOND:
                precond_items.append((col_d, set()))
            elif section == S_PARAM and current_param:
                param_data.setdefault(current_param, {})[-1] = col_d  # special: description for all

        # ── Result section rows with O markers ─────────────────────
        elif tcs and section in (S_CONFIRM, S_RETURN, S_EXCEPT, S_LOG):
            if section == S_CONFIRM:
                confirm_items.append((col_d, set(tcs)))
            elif section == S_RETURN:
                return_items.append((col_d, set(tcs)))
            elif section == S_EXCEPT:
                except_items.append((col_d, set(tcs)))
            elif section == S_LOG:
                log_items.append((col_d, set(tcs)))

    # ── Build markdown ─────────────────────────────────────────────
    emit(f"## F{fn:03d} — {func_name}")
    emit("")
    emit(f"**Function Code:** `{code}`  ")
    emit(f"**Created By:** {created_by}  ")
    emit(f"**Executed By:** {executed_by}  ")
    emit(f"**Lines of Code:** {loc}  ")
    emit("")
    emit(f"**Test Requirement:** {test_req}")
    emit("")
    emit("### Test Cases")
    emit("")
    emit("| # | Type | Test Description | Expected Result |")
    emit("|---|---|---|---|")

    for tc_idx, utcd_id in enumerate(utcd_ids):
        tc_type = tc_types.get(tc_idx, "—")

        # Test description: preconditions + parameter values
        desc_parts = []
        # Anonymous preconditions (col_C="None" rows without O)
        for desc, tcs_set in precond_items:
            if tc_idx in tcs_set:
                desc_parts.append(desc)
        # Named parameter values
        for param, vals in param_data.items():
            if tc_idx in vals:
                desc_parts.append(f"**{param}:** `{vals[tc_idx]}`")

        desc = "<br>".join(desc_parts) if desc_parts else "—"

        res_parts = []
        for label, bucket in [
            ("Confirm",    confirm_items),
            ("Return",     return_items),
            ("Exception",  except_items),
            ("Log",        log_items),
        ]:
            for d, tcs_set in bucket:
                if tc_idx in tcs_set:
                    res_parts.append(f"**{label}:** {d}")

        res = "<br>".join(res_parts) if res_parts else "—"

        emit(f"| `{utcd_id}` | {tc_type} | {desc} | {res} |")

    emit("")


def write_header():
    emit("# Unit Test Document — EDUACAS Backend")
    emit("")
    emit("| Property | Value |")
    emit("|---|---|")
    emit("| **Project Name** | EDUACAS Backend |")
    emit("| **Project Code** | EDUACAS |")
    emit("| **Creator** | QA Team |")
    emit("| **Issue Date** | 2026-04-23 |")
    emit("")
    emit("---")
    emit("")
    emit("## Table of Contents")
    emit("")

    func_names = []
    for fn in range(1, 27):
        ws = wb[f'Function {fn}']
        func_name = "Unknown"
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            # Collect all non-None cells; last one is the function name
            non_none = [c for c in row if c is not None]
            if non_none:
                func_name = str(non_none[-1]).strip()
        func_names.append((fn, func_name))
        anchor = func_name.lower()
        anchor = ''.join((c if c.isalnum() else '-') for c in anchor).strip('-')
        emit(f"- [F{fn:03d} — {func_name}](#f{fn:03d}---{anchor})")

    emit("")
    emit("---")
    emit("")

    return func_names


fn_names = write_header()

for fn, name in fn_names:
    parse_sheet(fn)
    emit("")

output_text = "\n".join(out_lines)
outpath = '/home/tienthuan29/workspaces/projects/Capstone-Sp26-EduACAS/documentations/testing/unit-tests-function1-26.md'
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(output_text)

print(f"Done. {len(output_text)} chars, {len(out_lines)} lines → {outpath}")
